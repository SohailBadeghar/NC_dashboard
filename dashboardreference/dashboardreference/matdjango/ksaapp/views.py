from django.shortcuts import render
#import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import io
import urllib, base64

from openpyxl import load_workbook
from .models import Product

def home(request):
    plt.plot(range(10))
    fig = plt.gcf()
    #convert graph into dtring buffer and then we convert 64 bit code into image
    buf = io.BytesIO()
    fig.savefig(buf,format='png')
    buf.seek(0)
    string = base64.b64encode(buf.read())
    uri =  urllib.parse.quote(string)
    return render(request,'home.html',{'data':uri})

def import_from_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            name, price, quantity = row
            Product.objects.create(name=name, price=price, quantity=quantity)
            print (name)

        return render(request, 'import_success.html')

    return render(request, 'import_form.html')